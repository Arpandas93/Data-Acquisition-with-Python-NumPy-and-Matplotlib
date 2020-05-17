import csv
import openpyxl
# w - write
# r - read
# a - append
# r+ - read + write

######Text File######
f = open('testfile.txt', 'w')
print(f)


f.write('first Line.\n')
f.write('second line.\n')
print(f)
f.close()

f = open('testfile.txt', 'r')
print(f.read())


######Csv File######
file = open('test.csv', 'w')
print(file)

file = open('test.csv', 'r')
print(file)

file.close()
csvfile = csv.reader(file, delimiter=',')
print(csvfile)

for row in csvfile:
    print(row)

file = open('test.csv', 'r')
print(file)
csvfile = csv.reader(file, delimiter=',')
print(csvfile)


for row in csvfile:
    for element in row:
        print(element)


file.close()

######Excel File######

wb = openpyxl.load_workbook('test.xlsx')
print(wb)
print(type(wb))

print(wb.sheetnames)
currSheet = wb['Sheet1']
print(currSheet)

currSheet = wb[wb.sheetnames[0]]
print(currSheet)
print(currSheet.title)

var1 = currSheet['A1']
print(var1.value)
print(currSheet['B1'].value)

var2 = currSheet.cell(row= 2, column=2)
print(var2.value)

print(currSheet.max_row)
print(currSheet.max_column)

for i in range(currSheet.max_row):
    print('--Biginning of Row')
    for j in range(currSheet.max_column):
        var = currSheet.cell(row=i+1, column= j+1)
        print(var.value)
    print('---End of Row---')


